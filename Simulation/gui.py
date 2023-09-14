import matplotlib.pyplot as plt
import math
from matplotlib.lines import Line2D
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
import numpy as np
import triangulation

# define global variables
x_mics = None
y_mics = None

x_test = None
y_test = None
x_res = None
y_res = None

est_toa = None
act_toa = None

act_tri = None
est_tri = None

x_tri = None
y_tri = None

noise = None

x_max = 0
y_max = 0
num_points = 0

parabolas = None

colours = None

test_dir = ""

mic_co_ords = None

freq = None

noise_t = ""


# plot test points
def plot_test_points():
    global x_test, y_test, x_max, y_max, test_dir

    # Create a scatter plot
    plt.scatter(x_test, y_test, color='blue', marker='o')

    # Add labels and a title
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title('Test Points on Grid')

    # Set limits for grid points
    plt.xlim(0, x_max)
    plt.ylim(0, y_max)

    # Save the figure as a JPEG file
    plt.savefig(test_dir + '/test_points.jpg', dpi=300)


# plot resultant points
def plot_result_points():
    global x_res, y_res, x_max, y_max, test_dir

    # Create a scatter plot
    plt.scatter(x_res, y_res, color='red', marker='o')

    # Add labels and a title
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title('Result Points on Grid')

    # Set limits for grid points
    plt.xlim(0, x_max)  # Set x-axis limits from 0 to 6
    plt.ylim(0, y_max)  # Set y-axis limits from 0 to 12

    # Save the figure as a JPEG file
    plt.savefig(test_dir + '/result_points.jpg', dpi=300)


# plot test and resultant points
def plot_test_result_points():
    global x_res, y_res, x_test, y_test, x_max, y_max, test_dir, colours
    colours = plt.cm.viridis(np.linspace(0, 1, len(x_test)))

    # Create a scatter plot for each pair of test and resultant points
    for i in range(len(x_test)):
        plt.scatter(
            x_test[i], y_test[i], label=f'Actual Point {i+1}', color=colours[i], marker='s', s=50)
        plt.scatter(
            x_res[i], y_res[i], label=f'Estimated Point {i+1}', color=colours[i], marker='o', s=50)

    legend_handles = [Line2D([0], [0], marker='s', color='w', label='Actual Points', markersize=10, markerfacecolor='black'), Line2D(
        [0], [0], marker='o', color='w', label='Estimated Points', markersize=10, markerfacecolor='black')]

    # Add labels and a legend
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title('Resultant & Test Points on Grid')
    plt.legend(handles=legend_handles)

    # Save the figure as a JPEG file
    plt.savefig(test_dir + '/test_result_points.jpg', dpi=300)


# plot test, result and parabolas for all array of test points
def plot_all_points():
    global x_test, y_test, x_res, y_res, x_mics, y_mics, x_max, y_max, parabolas, num_points

    # Create and save individual scatter plots
    for i, (x_t, y_t, x_r, y_r) in enumerate(zip(x_test, y_test, x_res, y_res)):
        plt.figure(figsize=(6, 4))
        for k in range(0, 3):
            plt.contour(x_tri, y_tri, parabolas[i][k], [0], colors=["black"])
        plt.scatter(
            x_t, y_t, label=f'Actual Point {i+1}', color=['cyan'], marker='o', s=150)
        plt.scatter(
            x_r, y_r, label=f'Estimated Point {i+1}', color=['red'], marker='.', s=150)
        plt.xlim(0, x_max)
        plt.ylim(0, y_max)
        plt.grid(True)
        plt.title(f"Test Point {i+1} Results")
        plt.xlabel("X")
        plt.ylabel("Y")
        plt.legend()
        plt.grid(True)
        plt.savefig(test_dir + f"/Point_{i+1}.jpg")
        plt.close()

    # Create and display all scatter plots together in one figure using subplots
    fig, axs = plt.subplots(2, 5, figsize=(20, 8))

    for i, (x_t, y_t, x_r, y_r, ax) in enumerate(zip(x_test, y_test, x_res, y_res, axs.ravel())):
        for k in range(0, 3):
            ax.contour(x_tri, y_tri, parabolas[i][k], [0], colors=["black"])
        ax.scatter(
            x_t, y_t, label=f'Actual Point {i+1}', color=['cyan'], marker='o', s=150)
        ax.scatter(
            x_r, y_r, label=f'Resultant Point {i+1}', color=['red'], marker='.', s=150)
        ax.set_title(f"Test Point {i+1}")
        ax.set_xlabel("X")
        ax.set_ylabel("Y")
        ax.legend()
        ax.grid(True)

    plt.tight_layout()

    # Save the final subplotted scatter plots as an image
    plt.savefig(test_dir + "/subplotted_point_results.jpg",
                bbox_inches="tight")

# generate xlsx for time of arrival results


def ss_toa_err():
    global x_max, est_toa, act_toa, num_points, noise

    points = np.linspace(1, num_points, num_points)

    err = np.empty((num_points, 3))
    p_f = np.array([True] * num_points, dtype=bool)

    if noise:
        acc = 3.0
    else:
        acc = 1.0

    for p in range(0, num_points):
        for m in range(0, 3):
            if act_toa[p][m] != 0.0:
                err[p][m] = str(
                    round((act_toa[p][m]-est_toa[p][m])/act_toa[p][m]*100, 5))
            else:
                err[p][m] = str(round(est_toa[p][m]/100, 5))
            if p_f[p] == True and err[p][m] and (err[p][m] > acc or err[p][m] < -1*acc):
                p_f[p] = False

    p_f_print = []
    for i in range(0, num_points):
        if p_f[i]:
            p_f_print.append("Pass")
        else:
            p_f_print.append("Fail")

    data = {
        'Points': points,
        'Mic Pair 1 Pecentage Error': err[:, 0],
        'Mic Pair 2 Pecentage Error': err[:, 1],
        'Mic Pair 3 Pecentage Error': err[:, 2],
        'Pass/Fail': p_f_print
    }

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Specify the Excel file name
    excel_file = test_dir + '/toa_err.xlsx'

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Convert the DataFrame to rows and add them to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    # Save the Excel file
    workbook.save(excel_file)


def ss_toa_val():
    global x_max, est_toa, act_toa, num_points

    points = np.linspace(1, num_points, num_points)

    # convert to correct format for displaying
    mic1_act = [row[0] for row in act_toa]
    mic2_act = [row[1] for row in act_toa]
    mic3_act = [row[2] for row in act_toa]
    mic1_est = [row[0] for row in est_toa]
    mic2_est = [row[1] for row in est_toa]
    mic3_est = [row[2] for row in est_toa]

    print(len(points))

    print(len(mic1_act))
    print(len(mic2_act))
    print(len(mic3_act))
    print(len(mic1_est))
    print(len(mic2_est))
    print(len(mic3_est))

    data = {
        'Points': points,
        'Mic Pair 1 Actual ': mic1_act,
        'Mic Pair 1 Estimated ': mic2_est,
        'Mic Pair 2 Actual ': mic3_act,
        'Mic Pair 2 Estimated ': mic1_est,
        'Mic Pair 3 Actual ': mic2_act,
        'Mic Pair 3 Estimated ': mic3_est,
    }

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Specify the Excel file name
    excel_file = test_dir + '/toa_cal.xlsx'

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Convert the DataFrame to rows and add them to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    # Apply number formatting to columns with 4 decimal places
    # for col in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
    #     for cell in col[1:]:
    #         cell.number_format = '0.0000'

    # Save the Excel file
    workbook.save(excel_file)

# generate xlsx for triangulation results


def ss_tri():
    global x_test, y_test, x_res, y_res, noise, x_max, y_max

    points = np.linspace(1, num_points, num_points)

    x_err = []
    y_err = []
    x_p_f = np.array([True] * num_points, dtype=bool)
    y_p_f = np.array([True] * num_points, dtype=bool)
    for i in range(0, num_points):
        x_err.append(round((x_test[i]-x_res[i])/x_max*100, 5))
        y_err.append(round((y_test[i]-y_res[i])/y_max*100, 5))
        if noise == True:
            acc = 5.0
        else:
            acc = 2.0
        if x_p_f[i] == True and (x_err[i] >= acc or x_err[i] <= -1 * acc):
            x_p_f[i] = False
        if y_p_f[i] == True and (y_err[i] >= acc or y_err[i] <= -1 * acc):
            y_p_f[i] = False

        x_err[i] = str(x_err[i])
        y_err[i] = str(y_err[i])

    p_f_print = []
    for i in range(0, num_points):
        if x_p_f[i] and y_p_f[i]:
            p_f_print.append("Pass")
        else:
            p_f_print.append("Fail")

    act = []
    est = []
    for i in range(0, num_points):
        act.append(
            '(' + str(round(x_test[i], 5)) + ',' + str(round(y_test[i], 5)) + ')')
        est.append(
            '(' + str(round(x_res[i], 5)) + ',' + str(round(y_res[i], 5)) + ')')
        # dist_err.append(
        #     math.sqrt((x_test[i] - x_res[i])**2 + (y_test[i] - y_res[i])**2))

    data = {
        'Points': points,
        'Actual Co-Ord': act,
        'Estimated Co-Ord': est,
        'X Percentage Error': x_err,
        'Y Percentage Error': y_err,
        'Pass/Fail': p_f_print
    }

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Specify the Excel file name
    excel_file = test_dir + '/tri_results.xlsx'

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Convert the DataFrame to rows and add them to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    # Save the Excel file
    workbook.save(excel_file)

# generate text file for simulation input parameteres (frequency/microphone points)


def sim_param_s():
    global test_dir, freq, mic_co_ords,noise_t

    file_name = test_dir + "/Test_Parameters.txt"
    with open(file_name, 'w') as file:
        # Write the data to the file
        file.write("Frequency: " + str(freq) + "\n")
        for i in range(1, 5):
            file.write(
                "Mic" + str(i) + ": (" + str(mic_co_ords[i-1][0]) + "," + str(mic_co_ords[i-1][1]) + ")\n")
        file.write("Noise Type: " + noise_t + "\n")


# create directories for test results


def create_dir():
    global test_dir

    # create main test results directory if not already made
    test_dir = "Test_Results"
    if not os.path.exists(test_dir):
        os.mkdir(test_dir)

    # create subsidiary test directory by finding the next available directory name
    found_dir = False
    count_dir = 1
    test_dir = test_dir + "/Test_Results"
    while found_dir == False:
        if not os.path.exists(test_dir + "_" + str(count_dir)):
            test_dir = test_dir + "_" + str(count_dir)
            os.mkdir(test_dir)
            found_dir = True
        count_dir += 1

# generate random parabola for testing


def rand_par(x_max):
    # Generate x values (e.g., from -5 to 5)
    x_values = np.linspace(0, x_max, 50)

    # Generate random coefficients for the parabola (a, b, c)
    a = np.random.uniform(-0.2, 0.2)
    b = np.random.uniform(-2.0, 2.0)
    c = np.random.uniform(0.0, 5.0)

    # Calculate y values using the parabola equation
    y_values = a * x_values**2 + b * x_values + c

    return x_values, y_values

# complete gui function for main sim program


def run(freq_in, mic_co_ords_in, x_test_in, y_test_in, x_res_in, y_res_in, x_max_in, y_max_in, num_points_in, est_toa_in, act_toa_in, parabolas_in, x_tri_in, y_tri_in, noise_in, noise_t_in):
    # set all necessary global variables
    global x_test, y_test, x_res, y_res, x_mics, y_mics, x_max, y_max, parabolas, num_points, est_toa, act_toa, x_tri, y_tri, freq, mic_co_ords, noise
    x_test = x_test_in
    y_test = y_test_in
    x_res = x_res_in
    y_res = y_res_in
    x_max = x_max_in
    y_max = y_max_in
    parabolas = parabolas_in
    num_points = num_points_in
    est_toa = est_toa_in
    act_toa = act_toa_in
    x_tri = x_tri_in
    y_tri = y_tri_in
    freq = freq_in
    mic_co_ords = mic_co_ords_in
    noise = noise_in
    noise_t = noise_t_in

    # create the relevant directories
    create_dir()

    # plot the result points
    plot_test_result_points()

    # plot test point, resultant point and parabolas for each test
    plot_all_points()

    # create xlsx file with toa results
    ss_toa_err()
    ss_toa_val()

    # create xlsx file with tri results
    ss_tri()

    # save system parameters to text file
    sim_param_s()

# main funtcion to test gui.py


def main():
    xs, ys = 40, 40
    x_test_in = np.array([xs, xs, xs, xs])
    y_test_in = np.array([ys, ys, ys, ys])
    x_max_in = 100
    y_max_in = 100
    num_points_in = 4
    freq_in = 280.0

    mic_co_ords_in = [0, 0], [0, 0.8], [0.5, 0], [0.5, 0.8]

    t1 = 40*np.sqrt(2)-20*np.sqrt(13)
    t2 = 40*np.sqrt(2)-60*np.sqrt(2)
    t3 = 40*np.sqrt(2)-20*np.sqrt(13)

    param = [0, 0, 0, 100, t1, 100, 100, t2, 100, 0, t3]
    xe, ye, x, y, h1, h2, h3 = triangulation.triangulate(
        param, [0, 100, 1, 0, 100, 1])
    parabolas = np.empty((num_points_in, 3), dtype=object)
    temp = [h1, h2, h3]
    for i in range(0, num_points_in):
        for k in range(0, 3):
            parabolas[i][k] = temp[k]

    x_res_in = [xe, xe, xe, xe]
    y_res_in = [ye, ye, ye, ye]

    # # est_toa_in = np.empty((num_points_in, 3))
    # # act_toa_in = np.empty((num_points_in, 3))
    # # for r in range(0, num_points_in):
    # #     for c in range(0, 3):
    # #         est_toa_in[r][c] = c
    # #         act_toa_in[r][c] = c

    # run(freq_in, mic_co_ords_in, x_test_in, y_test_in, x_res_in, y_res_in, x_max_in, y_max_in,
    #     num_points_in, est_toa_in, act_toa_in, parabolas, x, y, True)


# Check if the script is being run as the main program
if __name__ == "__main__":
    main()
